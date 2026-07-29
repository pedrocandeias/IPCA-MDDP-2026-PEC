#!/usr/bin/env python3
"""Aplica classificações Codex apenas às células vazias da matriz de evidência."""

from __future__ import annotations

import argparse
import hashlib
import json
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill


ALLOWED_LEVELS = {
    "Responde totalmente",
    "Responde parcialmente",
    "Não responde",
}


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def json_hash(payload: object) -> str:
    encoded = json.dumps(
        payload,
        ensure_ascii=False,
        separators=(",", ":"),
        default=str,
    ).encode("utf-8")
    return hashlib.sha256(encoded).hexdigest()


def default_rationale(level: str) -> str:
    if level == "Responde totalmente":
        return (
            "A fonte sustenta diretamente o núcleo factual e o alcance da "
            "afirmação a que está associada."
        )
    if level == "Responde parcialmente":
        return (
            "A fonte sustenta o núcleo temático ou uma parte factual da "
            "afirmação, mas não cobre todas as extensões, condições ou "
            "inferências formuladas no texto do DOCX."
        )
    return (
        "A fonte não sustenta a afirmação específica ou trata um objeto "
        "materialmente diferente."
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, required=True)
    parser.add_argument("--decisions", type=Path, required=True)
    parser.add_argument("--output", type=Path)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    workbook_path = args.workbook.resolve()
    decisions_path = args.decisions.resolve()
    output_path = args.output.resolve() if args.output else workbook_path

    if workbook_path.name == "rastreabilidade_citacoes.xlsx":
        raise SystemExit(
            "Este programa recusa explicitamente usar rastreabilidade_citacoes.xlsx."
        )

    decisions = json.loads(decisions_path.read_text(encoding="utf-8"))
    excluded = {Path(item).name for item in decisions.get("excluded_inputs", [])}
    if "rastreabilidade_citacoes.xlsx" not in excluded:
        raise ValueError(
            "O ficheiro de decisões não declara rastreabilidade_citacoes.xlsx "
            "como entrada excluída."
        )

    actual_workbook_hash = sha256_file(workbook_path)
    expected_workbook_hash = decisions["base_workbook_sha256"]
    workbook = load_workbook(workbook_path)
    sheet = workbook["Evidência por citação"]
    decision_marker = decisions_path.name
    codex_rows = [
        row
        for row in range(2, sheet.max_row + 1)
        if sheet.cell(row, 7).comment
        and decision_marker in sheet.cell(row, 7).comment.text
    ]
    if actual_workbook_hash != expected_workbook_hash and len(codex_rows) == 340:
        existing_rows = [
            row
            for row in range(2, sheet.max_row + 1)
            if row not in set(codex_rows)
        ]
        codex_rows_hash = hashlib.sha256(
            ",".join(map(str, codex_rows)).encode("utf-8")
        ).hexdigest()
        codex_content_hash = json_hash(
            [
                [row] + [sheet.cell(row, column).value for column in range(1, 7)]
                for row in codex_rows
            ]
        )
        existing_snapshot = [
            [
                row,
                sheet.cell(row, 7).value,
                sheet.cell(row, 7).comment.text
                if sheet.cell(row, 7).comment
                else None,
            ]
            for row in existing_rows
        ]
        final_counts = Counter(
            sheet.cell(row, 7).value
            for row in range(2, sheet.max_row + 1)
        )
        if (
            codex_rows_hash == decisions["expected_blank_rows_sha256"]
            and codex_content_hash == decisions["expected_blank_content_sha256"]
            and json_hash(existing_snapshot)
            == decisions["expected_existing_status_comment_sha256"]
            and final_counts.get(None, 0) == 0
        ):
            print("As 340 classificações Codex já estavam aplicadas e foram validadas.")
            print(f"Classificações preexistentes preservadas: {len(existing_rows)}")
            print(f"Totais finais: {dict(final_counts)}")
            print(workbook_path)
            return 0
        raise ValueError(
            "A matriz contém classificações Codex, mas não corresponde ao "
            "registo independente validado."
        )
    if actual_workbook_hash != expected_workbook_hash:
        raise ValueError(
            "A matriz de base não corresponde à versão revista: "
            f"{actual_workbook_hash}; esperado {expected_workbook_hash}."
        )

    blank_rows = [
        row
        for row in range(2, sheet.max_row + 1)
        if sheet.cell(row, 7).value in (None, "")
    ]
    existing_rows = [
        row
        for row in range(2, sheet.max_row + 1)
        if sheet.cell(row, 7).value not in (None, "")
    ]
    blank_rows_hash = hashlib.sha256(
        ",".join(map(str, blank_rows)).encode("utf-8")
    ).hexdigest()
    blank_content_hash = json_hash(
        [
            [row] + [sheet.cell(row, column).value for column in range(1, 7)]
            for row in blank_rows
        ]
    )
    existing_snapshot = [
        [
            row,
            sheet.cell(row, 7).value,
            sheet.cell(row, 7).comment.text
            if sheet.cell(row, 7).comment
            else None,
        ]
        for row in existing_rows
    ]
    existing_hash = json_hash(existing_snapshot)

    expected_checks = {
        "existing_count": (
            len(existing_rows),
            int(decisions["expected_existing_count"]),
        ),
        "existing_status_comment_sha256": (
            existing_hash,
            decisions["expected_existing_status_comment_sha256"],
        ),
        "blank_count": (
            len(blank_rows),
            int(decisions["expected_blank_count"]),
        ),
        "blank_rows_sha256": (
            blank_rows_hash,
            decisions["expected_blank_rows_sha256"],
        ),
        "blank_content_sha256": (
            blank_content_hash,
            decisions["expected_blank_content_sha256"],
        ),
    }
    failed = {
        key: values
        for key, values in expected_checks.items()
        if values[0] != values[1]
    }
    if failed:
        raise ValueError(f"A base da revisão independente mudou: {failed}")

    default_level = decisions["default_level"]
    if default_level not in ALLOWED_LEVELS:
        raise ValueError(f"Classificação predefinida inválida: {default_level}")
    partial_rows = {int(row) for row in decisions["partial_rows"]}
    no_response_rows = {int(row) for row in decisions["no_response_rows"]}
    blank_set = set(blank_rows)
    if partial_rows & no_response_rows:
        raise ValueError("Há linhas simultaneamente parciais e sem resposta.")
    if not (partial_rows | no_response_rows) <= blank_set:
        invalid = sorted((partial_rows | no_response_rows) - blank_set)
        raise ValueError(f"As decisões incluem linhas que não estavam vazias: {invalid}")

    notes = {
        int(row): str(note)
        for row, note in decisions.get("notes", {}).items()
    }
    fills = {
        "Responde totalmente": PatternFill("solid", fgColor="C6EFCE"),
        "Responde parcialmente": PatternFill("solid", fgColor="FFE699"),
        "Não responde": PatternFill("solid", fgColor="F4CCCC"),
    }

    control = workbook["Controlo técnico"]
    if control.max_column < 15:
        control.cell(1, 15).value = "Justificação da classificação"

    for row in blank_rows:
        if row in no_response_rows:
            level = "Não responde"
        elif row in partial_rows:
            level = "Responde parcialmente"
        else:
            level = default_level
        rationale = notes.get(row, default_rationale(level))
        cell = sheet.cell(row, 7)
        cell.value = level
        cell.fill = fills[level]
        cell.comment = Comment(
            "Classificação independente produzida por Codex.\n"
            "Entrada explicitamente não consultada: "
            "rastreabilidade_citacoes.xlsx.\n"
            f"Origem: {decisions_path.name}, linha XLSX {row}.\n"
            f"Justificação: {rationale}",
            "Codex",
        )
        control.cell(row, 12).value = level
        control.cell(row, 13).value = (
            f"{decisions_path.name}, linha XLSX {row}"
        )
        control.cell(row, 14).value = None
        control.cell(row, 15).value = rationale

    current_existing_snapshot = [
        [
            row,
            sheet.cell(row, 7).value,
            sheet.cell(row, 7).comment.text
            if sheet.cell(row, 7).comment
            else None,
        ]
        for row in existing_rows
    ]
    if current_existing_snapshot != existing_snapshot:
        raise AssertionError("Uma das 77 classificações preexistentes foi alterada.")

    methodology = workbook["Metodologia"]
    for row in range(1, methodology.max_row + 1):
        if methodology.cell(row, 1).value == "Grau de resposta":
            methodology.cell(row, 2).value = (
                "Classificação crítica da relação entre o paper e o texto do "
                "DOCX: «Responde totalmente», «Responde parcialmente» ou "
                "«Não responde». As 417 ocorrências estão classificadas; a "
                "lista suspensa limita os valores permitidos."
            )
    methodology.append(
        [
            "Classificações independentes Codex",
            (
                "340 ocorrências anteriormente vazias classificadas sem "
                "consulta a rastreabilidade_citacoes.xlsx; decisões registadas "
                f"em {decisions_path.name}."
            ),
        ]
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    verified = load_workbook(output_path)
    verified_sheet = verified["Evidência por citação"]
    final_counts = Counter(
        verified_sheet.cell(row, 7).value
        for row in range(2, verified_sheet.max_row + 1)
    )
    final_existing_snapshot = [
        [
            row,
            verified_sheet.cell(row, 7).value,
            verified_sheet.cell(row, 7).comment.text
            if verified_sheet.cell(row, 7).comment
            else None,
        ]
        for row in existing_rows
    ]
    if final_existing_snapshot != existing_snapshot:
        raise AssertionError(
            "A gravação alterou uma das 77 classificações preexistentes."
        )
    if sum(final_counts.values()) != 417 or final_counts.get(None, 0):
        raise AssertionError(f"Classificações finais incompletas: {final_counts}")

    print(f"Classificações Codex aplicadas: {len(blank_rows)}")
    print(f"Classificações preexistentes preservadas: {len(existing_rows)}")
    print(f"Totais finais: {dict(final_counts)}")
    print(output_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
