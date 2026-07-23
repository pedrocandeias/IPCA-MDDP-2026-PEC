#!/usr/bin/env python3
"""Generate an occurrence-level citation evidence workbook from the thesis.

The source DOCX is treated as immutable. Page numbers refer to the one-based PDF
pages produced from that DOCX snapshot. Evidence excerpts are recovered from
the locally consolidated PDFs and always retain their original language.
"""

from __future__ import annotations

import argparse
import json
import math
import re
import unicodedata
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path

import fitz
import numpy as np
from lxml import etree
from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sklearn.feature_extraction.text import TfidfVectorizer


WORD_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
NS = {"w": WORD_NS}
YEAR_RE = r"(?:19|20)\d{2}[a-z]?"

# Explicit aliases are limited to orthographic or known year inconsistencies in
# the frozen manuscript. The target is (first-author fragment, bibliography year).
REFERENCE_OVERRIDES: dict[tuple[str, str], tuple[str, str]] = {
    ("arrieta", "2019"): ("barredo arrieta", "2020"),
    ("cameron brooks", "2026"): ("brooks", "2026"),
    ("design council", "2007"): ("design council", "2020"),
    ("elhadad", "2025"): ("elhadad", "2026"),
    ("filho", "2023"): ("anacleto filho", "2023"),
    ("frangos", "2019"): ("frangos", "2016"),
    ("kate", "2017"): ("ten kate", "2017"),
    ("kuhl", "2021"): ("kuhl", "2020"),
    ("marinelli", "2023"): ("marinelli", "2022"),
    ("parlamento europeu do conselho europeu", "2017"): ("parlamento europeu", "2017"),
    ("resnik", "2022"): ("resnik", "2010"),
    ("shannon", "2019"): ("kellam", "2019"),
    ("showcase", "2007"): ("zimmerman", "2007"),
    ("thorsen", "2024"): ("thorsen", "2023"),
    ("van niekerk", "2021"): ("van niekerk", "2018"),
    ("ostlie", "2011"): ("ostlie", "2012"),
    ("alcara da silva", "sem data"): ("silva", "2018"),
}

FORCE_UNRESOLVED = {
    ("dickinson", "2024"),
}

PORTUGUESE_TO_ENGLISH = {
    "abandono": "abandonment rejection discontinuation",
    "aceitacao": "acceptance adoption",
    "acessibilidade": "accessibility accessible",
    "ajuste": "fit fitting adjustment socket fit",
    "amputacao": "amputation limb loss amputee",
    "antropometria": "anthropometry anthropometric",
    "antropometrico": "anthropometric",
    "aprendizagem": "learning training",
    "arnes": "harness cable suspension",
    "articulacao": "joint articulation",
    "atividade diaria": "activities of daily living daily activities ADL",
    "autonomia": "independence autonomy",
    "avaliacao": "evaluation assessment validation",
    "baixo custo": "low cost affordable",
    "bateria": "battery power supply",
    "clinico": "clinical clinician",
    "co-criacao": "co-creation cocreation",
    "co-design": "co-design codesign participatory design",
    "conforto": "comfort comfortable",
    "configuracao": "configuration configurator customization",
    "controlo": "control",
    "cosmetica": "cosmetic restorative passive",
    "custo": "cost affordability",
    "dados": "data dataset",
    "decisao": "decision decision-making",
    "dedo": "finger digit",
    "design inclusivo": "inclusive design design for all",
    "design participativo": "participatory design co-design",
    "dispositivo medico": "medical device healthcare device",
    "eficacia": "efficacy effectiveness outcome",
    "encaixe": "socket fit fitting",
    "erro": "error failure",
    "fabrico aditivo": "additive manufacturing 3D printing",
    "funcao": "function functionality functional",
    "geometria": "geometry geometric",
    "hibrida": "hybrid prosthesis body-powered elbow myoelectric hand",
    "humidade": "moisture water humidity",
    "impressao 3d": "3D printing three-dimensional printed",
    "inteligencia artificial": "artificial intelligence AI machine learning",
    "interface": "interface user interface usability",
    "mao": "hand",
    "manutencao": "maintenance repair service",
    "mecanica": "mechanical body-powered cable-operated",
    "medida": "measurement dimension",
    "membro inferior": "lower limb lower-limb",
    "membro superior": "upper limb upper-limb",
    "mioeletrica": "myoelectric externally powered electric prosthesis EMG",
    "modelacao parametrica": "parametric modelling parametric modeling",
    "necessidades": "needs requirements priorities",
    "open source": "open source open-source",
    "parametro": "parameter parametric",
    "participacao": "participation involvement engagement",
    "passiva": "passive cosmetic restorative prosthesis",
    "personalizacao": "personalization personalisation customization individualization",
    "peso": "weight lightweight heavy",
    "plataforma": "platform system workflow",
    "proprioceptiva": "proprioceptive proprioception cable feedback",
    "prototipo": "prototype prototyping",
    "protese": "prosthesis prosthetic",
    "psicossocial": "psychosocial social psychological",
    "reabilitacao": "rehabilitation",
    "responsabilidade": "responsibility accountability",
    "seguranca": "safety risk",
    "sensorial": "sensory feedback sensation",
    "sistema": "system framework",
    "supervisao humana": "human oversight human-in-the-loop",
    "tecnico": "technical",
    "transumeral": "transhumeral above-elbow",
    "treino": "training rehabilitation learning",
    "usabilidade": "usability user experience",
    "utilizador": "user end-user wearer patient",
    "validacao": "validation validated",
    "volume residual": "residual limb volume",
}


@dataclass
class BibliographyEntry:
    index: int
    raw: str
    authors: str
    year: str
    title: str
    first_author: str
    pdf: Path | None = None


@dataclass
class CitationOccurrence:
    page: int
    order: int
    literal: str
    label: str
    cited_year: str
    statement: str
    entry: BibliographyEntry | None = None
    match_score: float = 0.0
    evidence: str = ""
    evidence_page: int | None = None
    evidence_score: float = 0.0
    evidence_method: str = "recuperação lexical bilingue"


def normalise(value: str) -> str:
    value = (
        value.replace("ı", "i")
        .replace("ø", "o")
        .replace("Ø", "O")
        .replace("ł", "l")
        .replace("Ł", "L")
        .replace("æ", "ae")
        .replace("Æ", "AE")
        .replace("–", "-")
        .replace("—", "-")
    )
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    value = value.casefold()
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def compact(value: str) -> str:
    return normalise(value).replace(" ", "")


def clean_pdf_text(value: str) -> str:
    value = ILLEGAL_CHARACTERS_RE.sub("", value)
    value = value.replace("\u00ad", "")
    value = re.sub(r"(?<=\w)-\s*\n\s*(?=\w)", "", value)
    value = re.sub(r"\s*\n\s*", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def docx_paragraphs(path: Path) -> list[tuple[str, str]]:
    with zipfile.ZipFile(path) as archive:
        root = etree.fromstring(archive.read("word/document.xml"))
    paragraphs: list[tuple[str, str]] = []
    for paragraph in root.xpath(".//w:body//w:p", namespaces=NS):
        text = "".join(paragraph.xpath(".//w:t/text()", namespaces=NS))
        text = re.sub(r"\s+", " ", text).strip()
        styles = paragraph.xpath("./w:pPr/w:pStyle/@w:val", namespaces=NS)
        paragraphs.append((text, styles[0] if styles else ""))
    return paragraphs


def parse_bibliography(path: Path) -> list[BibliographyEntry]:
    entries: list[BibliographyEntry] = []
    for text, style in docx_paragraphs(path):
        if style != "Bibliografia" or not text:
            continue
        match = re.match(
            r"^(.+?)\s+\((?P<year>(?:19|20)\d{2}|n\.d\.)(?:,[^)]*)?\)\.\s+(?P<rest>.+)$",
            text,
            re.I,
        )
        if not match:
            continue
        authors = match.group(1)
        year = match.group("year")
        remainder = match.group("rest")
        first_author = authors.split(",", 1)[0].strip().rstrip(".")
        title_match = re.match(r"(.+?)(?:\.\s+(?=[A-ZÀ-ÖØ-Þ*])|$)", remainder)
        title = (title_match.group(1) if title_match else remainder).strip().strip("*")
        entries.append(
            BibliographyEntry(
                index=len(entries) + 1,
                raw=text,
                authors=authors.strip(),
                year=year,
                title=title,
                first_author=first_author,
            )
        )
    return entries


def load_manifest(path: Path) -> list[dict[str, str]]:
    return json.loads(path.read_text(encoding="utf-8"))


def manifest_reference_parts(value: str) -> tuple[str, str]:
    match = re.match(r"^(.+?)\s+\(((?:19|20)\d{2})\)$", value.strip())
    return match.groups() if match else (value, "")


def attach_pdfs(entries: list[BibliographyEntry], manifest: list[dict[str, str]], root: Path) -> None:
    unused = set(range(len(manifest)))
    for entry in entries:
        best: tuple[float, int] | None = None
        for index in unused:
            item = manifest[index]
            manifest_author, manifest_year = manifest_reference_parts(item["reference"])
            entry_raw_norm = normalise(entry.raw)
            manifest_title_norm = normalise(item["title"])
            if manifest_title_norm and manifest_title_norm in entry_raw_norm:
                title_score = 1.0
            else:
                title_score = SequenceMatcher(
                    None, normalise(entry.title), manifest_title_norm
                ).ratio()
            author_score = SequenceMatcher(
                None, normalise(entry.first_author), normalise(manifest_author)
            ).ratio()
            year_score = 1.0 if entry.year == manifest_year else 0.0
            score = 0.70 * title_score + 0.20 * author_score + 0.10 * year_score
            if year_score == 1.0 and author_score >= 0.82:
                score = max(score, 0.90)
            if best is None or score > best[0]:
                best = (score, index)
        if best and best[0] >= 0.58:
            item = manifest[best[1]]
            pdf_path = Path(item["pdf"])
            if not pdf_path.is_absolute():
                pdf_path = root / pdf_path
            if pdf_path.exists():
                entry.pdf = pdf_path.resolve()
                entry.title = item["title"].strip()
                unused.remove(best[1])


def append_unused_manifest_entries(
    entries: list[BibliographyEntry], manifest: list[dict[str, str]], root: Path
) -> None:
    """Expose validated legacy matches for citations missing from the DOCX bibliography."""
    used_pdfs = {entry.pdf for entry in entries if entry.pdf}
    for item in manifest:
        pdf = Path(item["pdf"])
        if not pdf.is_absolute():
            pdf = (root / pdf).resolve()
        if pdf in used_pdfs or not pdf.exists():
            continue
        author, year = manifest_reference_parts(item["reference"])
        entries.append(
            BibliographyEntry(
                index=len(entries) + 1,
                raw=f"{item['reference']}. {item['title']}",
                authors=author.rstrip("."),
                year=year,
                title=item["title"].strip(),
                first_author=author.rstrip("."),
                pdf=pdf,
            )
        )


def extract_pages(pdf_path: Path) -> list[str]:
    document = fitz.open(pdf_path)
    pages: list[str] = []
    for page in document:
        pages.append(page.get_text("text", sort=True))
    document.close()
    return pages


def manuscript_ranges(pages: list[str]) -> tuple[int, int, int]:
    intro = bibliography = annex = None
    for page_number, text in enumerate(pages, 1):
        if intro is None and re.search(r"(?m)^\s*1\s+Introdução\s*$", text):
            intro = page_number
        if bibliography is None and re.search(r"(?mi)^\s*Referências Bibliográficas\s*$", text):
            bibliography = page_number
        if annex is None and re.search(r"(?m)^\s*Anexo A\s+[—-]", text):
            annex = page_number
    if not all((intro, bibliography, annex)):
        raise RuntimeError("Não foi possível localizar introdução, bibliografia e Anexo A no PDF.")
    return int(intro), int(bibliography), int(annex)


def statement_around(text: str, start: int, end: int) -> str:
    protected = text
    for abbreviation in ("et al.", "p. ex.", "i.e.", "e.g.", "Fig.", "Eq."):
        protected = protected.replace(abbreviation, abbreviation.replace(".", "∯"))
    boundaries = [match.end() for match in re.finditer(r"[.!?](?:\s+|$)", protected[:start])]
    # Retain the cited sentence plus at most the two preceding sentences. The
    # hard distance cap prevents a short page or table from pulling in a whole
    # unrelated section when punctuation is sparse.
    left = boundaries[-3] if len(boundaries) >= 3 else 0
    left = max(left, start - 1200)
    right_match = re.search(r"[.!?](?:\s+|$)", protected[end:])
    right = end + right_match.end() if right_match else min(len(text), end + 450)
    result = protected[left:right].replace("∯", ".").strip(" |\n\t")
    if len(result) > 1500:
        result = result[max(0, start - left - 1000) : start - left + 500]
    return re.sub(r"\s+", " ", result).strip()


def clean_citation_label(value: str) -> str:
    value = re.sub(r"\bet\s+al\.?\b", "", value, flags=re.I)
    value = re.sub(r"\b(?:and|e)\b", " ", value, flags=re.I)
    value = value.replace("&", " ")
    tokens = [token for token in normalise(value).split() if len(token) > 1]
    return " ".join(tokens)


def citation_parts(part: str) -> tuple[str, str] | None:
    part = part.strip()
    if not part or "http" in part.casefold() or "doi.org" in part.casefold():
        return None
    if re.search(r"\b(?:consultad[ao]s?|março|julho|equivalente)\b", part, re.I):
        return None
    match = re.match(
        rf"^(?P<label>.+?)(?:,\s*|\s+)(?P<year>{YEAR_RE}|sem\s+data|s\.d\.)(?:\b|$)",
        part,
        re.I,
    )
    if not match:
        return None
    label = match.group("label").strip(" ,")
    if not re.search(r"[A-Za-zÀ-ÖØ-öø-ÿ]", label):
        return None
    return label, match.group("year")


def extract_occurrences(pages: list[str]) -> list[CitationOccurrence]:
    intro, bibliography, annex = manuscript_ranges(pages)
    included = set(range(intro, bibliography)) | set(range(annex, len(pages) + 1))
    occurrences: list[CitationOccurrence] = []
    for page_number in sorted(included):
        text = pages[page_number - 1]
        # Annex B ends with a references-only subsection; it is not running text.
        text = re.split(r"(?mi)^\s*B\.9\s+REFERÊNCIAS NORMATIVAS\s*$", text)[0]
        text = clean_pdf_text(text)
        parenthetical_spans: list[tuple[int, int]] = []
        caption_spans: list[tuple[int, int]] = []
        caption_pattern = re.compile(
            rf"\b(?:Adaptado|Reproduzido)\s+de\s+(?P<authors>[^()]{{2,420}}?)\s*"
            rf"\((?P<year>{YEAR_RE})\)",
            re.I,
        )
        for caption in caption_pattern.finditer(text):
            authors = caption.group("authors").strip()
            if authors.startswith("(") or "http" in authors.casefold():
                continue
            first_author_match = re.search(r"([A-Za-zÀ-ÖØ-öø-ÿ’'\-]+)\s*,", authors)
            if not first_author_match:
                continue
            label = first_author_match.group(1)
            caption_spans.append(caption.span())
            occurrences.append(
                CitationOccurrence(
                    page=page_number,
                    order=caption.start(),
                    literal=caption.group(0),
                    label=label,
                    cited_year=caption.group("year"),
                    statement=statement_around(text, caption.start(), caption.end()),
                )
            )
        parenthetical_pattern = re.compile(
            rf"\([^()]{{0,360}}(?:{YEAR_RE}|sem\s+data|s\.d\.)[^()]{{0,180}}\)", re.I
        )
        for group in parenthetical_pattern.finditer(text):
            if any(left <= group.start() < right for left, right in caption_spans):
                continue
            if re.search(r"\.{5,}", text[max(0, group.start() - 220) : group.end() + 220]):
                continue
            parenthetical_spans.append(group.span())
            inner = group.group(0)[1:-1]
            for part in inner.split(";"):
                parsed = citation_parts(part)
                if not parsed:
                    continue
                label, year = parsed
                occurrences.append(
                    CitationOccurrence(
                        page=page_number,
                        order=group.start(),
                        literal=part.strip(),
                        label=label,
                        cited_year=year,
                        statement=statement_around(text, group.start(), group.end()),
                    )
                )

        narrative_pattern = re.compile(
            rf"(?P<label>(?:[A-ZÀ-ÖØ-Þ][\wÀ-ÿ’'\-.]+(?:\s+[A-ZÀ-ÖØ-Þ][\wÀ-ÿ’'\-.]+){{0,3}})"
            rf"(?:\s+(?:et\s+al\.|&\s+[A-ZÀ-ÖØ-Þ][\wÀ-ÿ’'\-.]+|e\s+[A-ZÀ-ÖØ-Þ][\wÀ-ÿ’'\-.]+))?)"
            rf"\s*\((?P<year>{YEAR_RE})\)",
        )
        for match in narrative_pattern.finditer(text):
            if any(left <= match.start() < right for left, right in parenthetical_spans + caption_spans):
                continue
            if re.search(r"\.{5,}", text[max(0, match.start() - 220) : match.end() + 220]):
                continue
            label = match.group("label").strip()
            # Avoid ordinary phrases ending in a capitalised word before a year.
            if len(label.split()) > 6:
                continue
            occurrences.append(
                CitationOccurrence(
                    page=page_number,
                    order=match.start(),
                    literal=match.group(0),
                    label=label,
                    cited_year=match.group("year"),
                    statement=statement_around(text, match.start(), match.end()),
                )
            )

    unique: dict[tuple[int, int, str, str], CitationOccurrence] = {}
    for occurrence in occurrences:
        key = (
            occurrence.page,
            occurrence.order,
            normalise(occurrence.label),
            occurrence.cited_year.casefold(),
        )
        unique.setdefault(key, occurrence)
    return sorted(unique.values(), key=lambda item: (item.page, item.order, item.literal))


def reference_similarity(label: str, cited_year: str, entry: BibliographyEntry) -> float:
    cleaned = clean_citation_label(label)
    first = normalise(entry.first_author)
    full_first_compact = compact(entry.first_author)
    label_compact = compact(cleaned)
    label_tokens = cleaned.split()
    first_tokens = first.split()
    ratio = SequenceMatcher(None, label_compact, full_first_compact).ratio()
    if (
        label_compact
        and min(len(label_compact), len(full_first_compact)) >= 4
        and (label_compact in full_first_compact or full_first_compact in label_compact)
    ):
        ratio = max(ratio, 0.94)
    if label_tokens and first_tokens and label_tokens[-1] == first_tokens[-1]:
        ratio = max(ratio, 0.90)
    if label_tokens and first_tokens and label_tokens[-1] in first_tokens:
        ratio = max(ratio, 0.84)
    if label_tokens and first_tokens and first_tokens[-1] in label_tokens:
        ratio = max(ratio, 0.90)
    base_year = re.match(r"\d{4}", cited_year)
    if base_year and base_year.group(0) == entry.year:
        year_score = 1.0
    elif base_year and entry.year.isdigit() and abs(int(base_year.group(0)) - int(entry.year)) <= 1:
        year_score = 0.25
    elif cited_year.casefold().startswith(("sem", "s.d")):
        year_score = 0.0
    else:
        year_score = -0.15
    return 0.78 * ratio + 0.22 * year_score


def resolve_occurrences(
    occurrences: list[CitationOccurrence], entries: list[BibliographyEntry]
) -> None:
    for occurrence in occurrences:
        override_key = (clean_citation_label(occurrence.label), occurrence.cited_year.casefold())
        if override_key in FORCE_UNRESOLVED:
            continue
        override = REFERENCE_OVERRIDES.get(override_key)
        candidates: list[tuple[float, BibliographyEntry]] = []
        for entry in entries:
            if override:
                target_author, target_year = override
                if target_author not in normalise(entry.first_author) or entry.year != target_year:
                    continue
            score = reference_similarity(occurrence.label, occurrence.cited_year, entry)
            candidates.append((score, entry))
        candidates.sort(key=lambda pair: pair[0], reverse=True)
        if candidates and (override or candidates[0][0] >= 0.56):
            occurrence.match_score, occurrence.entry = candidates[0]


def split_sentences(value: str) -> list[str]:
    value = clean_pdf_text(value)
    if not value:
        return []
    protected = value
    for abbreviation in ("et al.", "e.g.", "i.e.", "Fig.", "Eq.", "Dr.", "Prof."):
        protected = protected.replace(abbreviation, abbreviation.replace(".", "∯"))
    parts = re.split(r"(?<=[.!?])\s+(?=[A-ZÀ-ÖØ-Þ0-9])", protected)
    return [part.replace("∯", ".").strip() for part in parts if part.strip()]


def noisy_paper_fragment(value: str) -> bool:
    lowered = value.casefold()
    if any(
        marker in lowered
        for marker in (
            "to cite this article",
            "open access article distributed",
            "creative commons attribution",
            "all rights reserved",
        )
    ):
        return True
    if value.startswith("©") or re.match(r"^keywords?\b", value, re.I):
        return True
    if re.match(r"^(?:fig(?:ure)?|table)\.?\s*\d+", value, re.I):
        return True
    if not re.search(r"[.!?](?:\s|$)", value) and len(value) < 500:
        return True
    if len(re.findall(r"(?:^|\s)\d+\.\s+[A-ZÀ-ÖØ-Þ]", value)) >= 2:
        return True
    if len(re.findall(r"\b(?:19|20)\d{2}\b", value)) >= 3 and len(
        re.findall(r"\bet\s+al\.?", value, re.I)
    ) >= 2:
        return True
    if ("doi:" in lowered or "https://doi.org" in lowered) and len(value) < 260:
        return True
    if "phd student" in lowered and len(value) < 400:
        return True
    if "received" in lowered and "accepted" in lowered and len(value) < 450:
        return True
    if re.search(r"\bpage\s+\d+\s+of\s+\d+\b", value, re.I) and len(value) < 240:
        return True
    words = re.findall(r"[A-Za-zÀ-ÖØ-öø-ÿ]{3,}", value)
    return len(words) < 10


def paper_candidates(pdf: Path) -> list[tuple[int, str]]:
    document = fitz.open(pdf)
    candidates: list[tuple[int, str]] = []
    metadata_subject = clean_pdf_text(document.metadata.get("subject", ""))
    if len(metadata_subject) >= 100:
        candidates.append((1, metadata_subject[:1100]))
    references_started = False
    for page_number, page in enumerate(document, 1):
        page_text = page.get_text("text", sort=True)
        if page_number > max(2, len(document) // 3) and re.search(
            r"(?mi)(?:^|\n)\s*(?:\d+\.?\s+)?(?:references|bibliography|referências)\s*(?:\n|$)",
            page_text,
        ):
            references_started = True
        if references_started:
            continue
        blocks = page.get_text("blocks", sort=True)
        page_fragments: list[str] = []
        for block in blocks:
            value = clean_pdf_text(str(block[4]))
            if len(value) < 35 or value.count("http") > 2:
                continue
            if noisy_paper_fragment(value):
                continue
            value_lower = value.casefold()
            if "to cite this article" in value_lower or re.search(r"\.{5,}", value):
                continue
            if value.startswith("©") and len(value) < 400:
                continue
            if value_lower.count("doi") >= 2 or len(re.findall(r"\[\d+\]", value)) >= 4:
                continue
            if len(re.findall(r"\bet\s+al\.,", value, re.I)) >= 3:
                continue
            if ("@" in value or "correspondence:" in value_lower) and len(value) < 400:
                continue
            if re.match(r"^(?:fig(?:ure)?|table)\.?\s*\d+", value, re.I) and len(value) < 220:
                continue
            if len(re.findall(r"\b(?:19|20)\d{2}\b", value)) > 7:
                continue
            sentences = split_sentences(value)
            if not sentences:
                continue
            if len(value) <= 650 and (len(value) >= 120 or len(sentences) >= 2):
                page_fragments.append(value)
            else:
                for index in range(len(sentences)):
                    chunk = " ".join(sentences[index : index + 2])
                    if 80 <= len(chunk) <= 650:
                        page_fragments.append(chunk)
        # Short blocks often represent neighbouring columns or list items.
        for index in range(len(page_fragments) - 1):
            merged = f"{page_fragments[index]} {page_fragments[index + 1]}"
            if 140 <= len(merged) <= 650:
                candidates.append((page_number, merged))
        candidates.extend(
            (page_number, fragment[:650])
            for fragment in page_fragments
            if 100 <= len(fragment) <= 650
        )
    document.close()
    deduplicated: dict[str, tuple[int, str]] = {}
    for page_number, value in candidates:
        key = normalise(value)[:500]
        deduplicated.setdefault(key, (page_number, value))
    return list(deduplicated.values())


def expand_query(statement: str) -> str:
    norm = normalise(statement)
    additions: list[str] = []
    for source, target in PORTUGUESE_TO_ENGLISH.items():
        if normalise(source) in norm:
            additions.append(target)
    return f"{statement} {' '.join(additions)}"


def rank_evidence(
    statement: str, candidates: list[tuple[int, str]], paper_title: str = ""
) -> tuple[int | None, str, float]:
    if not candidates:
        return None, "", 0.0
    texts = [value for _, value in candidates]
    sentence_parts = split_sentences(statement)
    cited_sentence = sentence_parts[-1] if sentence_parts else statement
    cited_sentence = re.sub(
        rf"\([^()]*(?:{YEAR_RE}|sem\s+data|s\.d\.)[^()]*\)", " ", cited_sentence, flags=re.I
    )
    query = expand_query(f"{cited_sentence} {cited_sentence} {statement} {paper_title}")
    corpus = texts + [query]
    try:
        word_vectorizer = TfidfVectorizer(
            strip_accents="unicode",
            lowercase=True,
            ngram_range=(1, 2),
            sublinear_tf=True,
            max_features=45000,
        )
        word_matrix = word_vectorizer.fit_transform(corpus)
        word_scores = (word_matrix[:-1] @ word_matrix[-1].T).toarray().ravel()
    except ValueError:
        word_scores = np.zeros(len(texts))
    try:
        char_vectorizer = TfidfVectorizer(
            strip_accents="unicode",
            lowercase=True,
            analyzer="char_wb",
            ngram_range=(3, 5),
            min_df=2,
            sublinear_tf=True,
            max_features=55000,
        )
        char_matrix = char_vectorizer.fit_transform(corpus)
        char_scores = (char_matrix[:-1] @ char_matrix[-1].T).toarray().ravel()
    except ValueError:
        char_scores = np.zeros(len(texts))
    scores = 0.68 * word_scores + 0.32 * char_scores
    best_index = int(np.argmax(scores))
    page, excerpt = candidates[best_index]
    sentences = split_sentences(excerpt)
    deduplicated: list[str] = []
    for sentence in sentences:
        if deduplicated and normalise(sentence) == normalise(deduplicated[-1]):
            continue
        deduplicated.append(sentence)
    excerpt = " ".join(deduplicated) if deduplicated else excerpt
    return page, excerpt, float(scores[best_index])


def reviewed_visual_evidence(occurrence: CitationOccurrence) -> tuple[int, str] | None:
    """Short transcriptions for image-only PDFs reviewed page by page."""
    if not occurrence.entry:
        return None
    author = normalise(occurrence.entry.first_author)
    statement = normalise(occurrence.statement)
    if author == "chapman" and occurrence.entry.year == "2025":
        if "prisma" in statement or "identificacao e selecao" in statement:
            return (
                2,
                "The scoping review was guided by the first five stages of Arksey and O’Malley’s (2005) framework for scoping literature reviews, with Levac et al. (2010) enhancements.",
            )
        if "transparencia" in statement or "rigor" in statement:
            return (
                1,
                "Improved transparency and consistency in reporting co-design processes are recommended to enhance rigor and effectiveness of future initiatives.",
            )
        if "experiencias" in statement or "tomada de decisao" in statement or "poder de decisao" in statement:
            return (
                2,
                "Importantly, people with disability should be involved in shaping and designing the outcomes to reflect their needs.",
            )
        return (
            1,
            "Co-design, a collaborative approach where end-users are actively involved in design processes, has gained traction in health communication.",
        )
    if author == "frayling" and occurrence.entry.year == "1994":
        return (
            8,
            "Research through art and design [...] is less straightforward, but still identifiable and visible. [...] The thinking is, so to speak, embodied in the artefact.",
        )
    if author == "guo" and occurrence.entry.year == "2025":
        return (
            1,
            "Integrating theoretical modeling with empirical investigation, the research identifies core design determinants aligned with user expectations. [...] The study explores Chinese prosthetic users’ functional, emotional, and social integration needs.",
        )
    return None


def attach_evidence(occurrences: list[CitationOccurrence]) -> None:
    by_pdf: dict[Path, list[CitationOccurrence]] = defaultdict(list)
    for occurrence in occurrences:
        if occurrence.entry and occurrence.entry.pdf:
            by_pdf[occurrence.entry.pdf].append(occurrence)
    for pdf, relevant in sorted(by_pdf.items(), key=lambda pair: str(pair[0])):
        candidates = paper_candidates(pdf)
        for occurrence in relevant:
            reviewed = reviewed_visual_evidence(occurrence)
            if reviewed:
                occurrence.evidence_page, occurrence.evidence = reviewed
                occurrence.evidence_score = 1.0
                occurrence.evidence_method = "transcrição após revisão visual do PDF"
                continue
            page, excerpt, score = rank_evidence(
                occurrence.statement,
                candidates,
                occurrence.entry.title if occurrence.entry else "",
            )
            occurrence.evidence_page = page
            occurrence.evidence = excerpt
            occurrence.evidence_score = score


def source_text(occurrence: CitationOccurrence) -> str:
    if occurrence.evidence and occurrence.evidence_page:
        return f"[PDF, p. {occurrence.evidence_page}] {occurrence.evidence}"
    if occurrence.entry and not occurrence.entry.pdf:
        return "[Sem texto integral local validado; não foi possível extrair um excerto.]"
    if occurrence.entry and occurrence.entry.pdf:
        return "[PDF local sem camada textual utilizável; exige OCR ou revisão visual.]"
    return "[Referência citada não localizada de forma inequívoca na bibliografia do DOCX.]"


def write_workbook(
    path: Path,
    occurrences: list[CitationOccurrence],
    docx: Path,
    pdf: Path,
    bibliography_count: int,
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Evidência por citação"
    headers = [
        "Página do nosso DOCX",
        "Texto do nosso DOCX referenciado",
        "Título do paper",
        "Autor do paper",
        "Ano de publicação do paper",
        "Texto extraído do paper para fundamentar a referência no nosso DOCX",
    ]
    sheet.append(headers)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    missing_fill = PatternFill("solid", fgColor="F4CCCC")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[1].height = 42

    for occurrence in occurrences:
        entry = occurrence.entry
        row = [
            occurrence.page,
            occurrence.statement,
            entry.title if entry else "[Não localizado na bibliografia do DOCX]",
            entry.authors if entry else occurrence.label,
            entry.year if entry else occurrence.cited_year,
            source_text(occurrence),
        ]
        sheet.append(row)
        row_number = sheet.max_row
        if entry and entry.pdf:
            relative = Path("..") / entry.pdf.name
            sheet.cell(row_number, 3).hyperlink = relative.as_posix()
            sheet.cell(row_number, 3).style = "Hyperlink"
            comment = (
                f"Citação no DOCX: {occurrence.literal}\n"
                f"PDF local: {entry.pdf.name}\n"
                f"Correspondência bibliográfica: {occurrence.match_score:.3f}\n"
                f"Método do excerto: {occurrence.evidence_method}\n"
                f"Pontuação de recuperação do excerto: {occurrence.evidence_score:.3f}"
            )
            sheet.cell(row_number, 6).comment = Comment(comment, "Codex")
            if occurrence.evidence_score < 0.05:
                sheet.cell(row_number, 6).fill = warning_fill
        else:
            sheet.cell(row_number, 6).fill = missing_fill
        for cell in sheet[row_number]:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = [20, 74, 54, 48, 22, 92]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.sheet_view.showGridLines = False

    method = workbook.create_sheet("Metodologia")
    method_rows = [
        ("Objeto", "Matriz de ocorrências de citações e evidência textual primária."),
        ("DOCX congelado", docx.name),
        ("PDF paginado", pdf.name),
        ("Paginação", "Número físico da página (base 1) no PDF gerado a partir da cópia congelada do DOCX."),
        ("Excerto do paper", "Texto na língua original; a indicação [PDF, p. n] usa a página física do PDF local."),
        ("Unidade de registo", "Uma linha por referência dentro de cada ocorrência; grupos com várias fontes originam várias linhas."),
        ("Contexto do DOCX", "A célula inclui a frase citante e, quando disponíveis na mesma página, até duas frases anteriores para preservar a progressão argumentativa."),
        ("Âmbito", "Corpo do manuscrito e anexos; índice, bibliografia final e listas de referências normativas foram excluídos."),
        ("Identificação", "Autoria, ano e título provêm da bibliografia do DOCX; o PDF é associado pelo manifesto local validado."),
        ("Recuperação", "Seleção lexical bilingue de um excerto do PDF relacionado com a frase citante; células amarelas indicam baixa pontuação e requerem revisão humana."),
        ("Ausências", "Células vermelhas assinalam texto integral local inexistente ou referência não resolvida; não foi criada evidência substituta."),
        ("Entradas bibliográficas", bibliography_count),
        ("Ocorrências registadas", len(occurrences)),
        ("Ocorrências com PDF e excerto", sum(bool(item.evidence) for item in occurrences)),
        ("Ocorrências sem PDF ou sem correspondência", sum(not bool(item.evidence) for item in occurrences)),
    ]
    for key, value in method_rows:
        method.append([key, value])
    for cell in method[1]:
        cell.font = Font(bold=True)
    method.column_dimensions["A"].width = 32
    method.column_dimensions["B"].width = 120
    for row in method.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    method.sheet_view.showGridLines = False

    control = workbook.create_sheet("Controlo técnico")
    control.sheet_state = "hidden"
    control.append(
        [
            "Página",
            "Ordem",
            "Citação literal",
            "Rótulo extraído",
            "Ano citado",
            "Referência resolvida",
            "Pontuação da correspondência",
            "PDF local",
            "Página do PDF do paper",
            "Pontuação do excerto",
            "Método do excerto",
        ]
    )
    for occurrence in occurrences:
        control.append(
            [
                occurrence.page,
                occurrence.order,
                occurrence.literal,
                occurrence.label,
                occurrence.cited_year,
                occurrence.entry.raw if occurrence.entry else "",
                occurrence.match_score,
                occurrence.entry.pdf.name if occurrence.entry and occurrence.entry.pdf else "",
                occurrence.evidence_page,
                occurrence.evidence_score,
                occurrence.evidence_method,
            ]
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--docx", type=Path, required=True)
    parser.add_argument("--pdf", type=Path, required=True)
    parser.add_argument("--manifest", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    docx = args.docx.resolve()
    pdf = args.pdf.resolve()
    manifest_path = args.manifest.resolve()
    output = args.output.resolve()
    root = manifest_path.parent.parent
    entries = parse_bibliography(docx)
    if not entries:
        raise SystemExit("Não foram encontradas entradas bibliográficas no DOCX.")
    bibliography_count = len(entries)
    manifest = load_manifest(manifest_path)
    attach_pdfs(entries, manifest, root)
    append_unused_manifest_entries(entries, manifest, root)
    pages = extract_pages(pdf)
    occurrences = extract_occurrences(pages)
    resolve_occurrences(occurrences, entries)
    attach_evidence(occurrences)
    write_workbook(output, occurrences, docx, pdf, bibliography_count)
    print(f"Bibliografia: {bibliography_count}")
    print(f"PDFs associados: {sum(bool(entry.pdf) for entry in entries)}")
    print(f"Ocorrências: {len(occurrences)}")
    print(f"Ocorrências resolvidas: {sum(bool(item.entry) for item in occurrences)}")
    print(f"Ocorrências com excerto: {sum(bool(item.evidence) for item in occurrences)}")
    print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
