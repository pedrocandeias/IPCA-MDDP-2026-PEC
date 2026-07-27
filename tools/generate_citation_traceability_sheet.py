#!/usr/bin/env python3
"""Gera a folha de rastreabilidade das citações do DOCX canónico.

Uma linha por par (ocorrência no texto, fonte citada), com as colunas:
página impressa do DOCX, texto citante, título, autores e ano da fonte,
excerto do PDF da fonte que melhor corresponde à afirmação, e o grau de
confiança do excerto (automático — carece de verificação humana).

Entradas: o DOCX canónico, o PDF regenerado para paginação, a região da
bibliografia do próprio DOCX e o mapa entrada→PDF de
`projecto_completo_bibliografia/copied_matches.json`.

Saída: docs/revisoes/rastreabilidade_citacoes.xlsx
"""

from __future__ import annotations

import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path
from zipfile import ZipFile

import fitz
from lxml import etree
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DOCX = ROOT / "pedro-candeias-projeto-mestrado-mdddp-ipca-2026-revisto.docx"
PDF = ROOT / "docs/revisoes/.tmp-pdf/pedro-candeias-projeto-mestrado-mdddp-ipca-2026-revisto.pdf"
MATCHES = ROOT / "projecto_completo_bibliografia/copied_matches.json"
OUT = ROOT / "projecto_completo_bibliografia/auditoria_citacoes/rastreabilidade_citacoes.xlsx"

PRELIM_OFFSET = 18   # páginas preliminares antes da página impressa 1

W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
PARTICLES = {"da", "de", "do", "dos", "das", "van", "von", "der", "la",
             "el", "ten", "ter"}
ORG_ALIASES = {"iso": "international organization for standardization",
               "iec": "international electrotechnical commission",
               "astm": "astm international"}

PART = r"(?:da|de|do|dos|das|van|von|der|ten|ter)"
NAME = (r"(?:" + PART + r"\s+)*[A-ZÀ-Þ][\w'’\-]+"
        r"(?:\s+(?:" + PART + r")\s+[A-ZÀ-Þ][\w'’\-]+)*")
ORG = r"(?:ISO|IEC|ASTM|OpenSCAD (?:Project|Community)|Design Council|" \
      r"Center for Universal Design|Parlamento Europeu(?:\s*&\s*Conselho[^,;]*)?)"
INITS = r"(?:[A-ZÀ-Þ]\.\s*(?:C\.\s*)?(?:da\s+S\.\s*)?)*"
FIRST = r"(?:" + INITS + NAME + r"|" + ORG + r")"
YEAR = r"(\d{4}[a-z]?|n\.d\.(?:-[a-z])?)"

SEGMENT = re.compile(
    r"^(?:e\.g\.,?\s*|cf\.\s*|ver\s+|vide\s+|p\.\s*ex\.,?\s*)?(" + FIRST + r")"
    r"(?:\s+et\s+al\.?|\s*(?:&|e)\s*(" + NAME + r"))?"
    r",\s*" + YEAR + r"(?:,\s*pp?\.\s*[\d\-–]+)?$")
NARRATIVE = re.compile(
    r"(" + FIRST + r")"
    r"(?:\s+et\s+al\.?|\s+(?:&|e)\s+(" + NAME + r"))?\s*"
    r"\(" + YEAR + r"\)")


def deaccent(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if not unicodedata.combining(c))


def key_name(surname: str) -> str:
    s = re.sub(r"[A-ZÀ-Þ]\.\s*", "", surname)      # remove iniciais
    toks = deaccent(s).lower().replace("’", "'").split()
    toks = [t for t in toks if t not in PARTICLES]
    return toks[-1] if toks else deaccent(surname).lower()


def initials_of(citation_first: str) -> str:
    return "".join(re.findall(r"([A-ZÀ-Þ])\.", citation_first))


def norm_txt(s: str) -> str:
    return re.sub(r"\s+", " ", unicodedata.normalize("NFKC", s)).strip()


def para_text(p) -> str:
    return "".join(n.text or "" for n in p.iter(W + "t"))


def parse_bib_entry(entry: str):
    m = re.search(r"\((\d{4}|n\.d\.)[^)]*\)\.?", entry)
    if not m:
        return None
    year = m.group(1)
    head = entry[:m.start()].strip().rstrip(".")
    rest = entry[m.end():].strip()
    tm = re.match(r"([^.]+(?:\.[^ .][^.]*)*?)[.?!](?:\s|$)", rest)
    title = tm.group(1).strip() if tm else rest[:120]
    surnames = re.findall(
        r"((?:(?:da|de|do|dos|das|van|von|der|ten|ter)\s+)*[A-ZÀ-Þ][\w'’\-]+"
        r"(?:\s+[A-ZÀ-Þ][\w'’\-]+)*),\s*(?:[A-ZÀ-Þ]\.\s*[‑\-]?\s*)+", head)
    if not surnames:
        surnames = [head.split(".")[0].split(",")[0].strip()]
    return {"authors": head, "year": year, "title": title,
            "surnames": surnames, "text": entry}


def sentences(text: str) -> list[str]:
    parts = re.split(r"(?<=[.!?])\s+(?=[A-ZÀ-Þ«(])", text)
    return [p.strip() for p in parts if len(p.strip()) > 30]


HAS_CITATION = re.compile(r"\([^()]*\b(?:19|20)\d{2}[a-z]?\b[^()]*\)|"
                          r"\b(?:19|20)\d{2}[a-z]?\)|n\.d\.")


def claim_sentence(text: str, pos: int) -> str:
    """Frase citante mais até duas frases anteriores sem citações próprias."""
    bounds = [0]
    for m in re.finditer(r"(?<=[.!?])\s+(?=[A-ZÀ-Þ«(])", text):
        bounds.append(m.end())
    bounds.append(len(text))
    spans = [(a, b) for a, b in zip(bounds, bounds[1:]) if a < b]
    cur = next((k for k, (a, b) in enumerate(spans) if a <= pos < b),
               len(spans) - 1)
    start_idx = cur
    for k in range(cur - 1, max(cur - 3, -1), -1):
        if HAS_CITATION.search(text[spans[k][0]:spans[k][1]]):
            break
        start_idx = k
    start = spans[start_idx][0]
    end = spans[cur][1]
    return norm_txt(text[start:end])[:800]


STOP = set("""a o os as um uma de do da dos das em no na nos nas por para com
sem sobre entre que se ao aos à às e ou não mais menos como the of and in on
for with to from this that are is was were their its d’ deste desta neste
nesta pelo pela pelos pelas""".split())


PT_EN = {
    "protese": ["prosthes", "prosthet"], "proteses": ["prosthes", "prosthet"],
    "mao": ["hand"], "maos": ["hands"], "membro": ["limb"],
    "impressao": ["print"], "impressas": ["printed"], "impressos": ["printed"],
    "utilizador": ["user"], "utilizadores": ["users"],
    "crianca": ["child"], "criancas": ["children"],
    "medida": ["measure"], "medidas": ["measure", "dimension"],
    "dedo": ["finger"], "dedos": ["finger", "digit"],
    "encaixe": ["socket"], "peso": ["weight"], "custo": ["cost"],
    "custos": ["cost"], "conforto": ["comfort"], "abandono": ["abandon", "rejection"],
    "rejeicao": ["rejection"], "aprendizagem": ["learning"],
    "desenho": ["design"], "avaliacao": ["evaluation", "assessment"],
    "ensaios": ["trial", "test"], "saude": ["health"],
    "doentes": ["patient"], "paciente": ["patient"],
    "fabrico": ["manufactur", "fabrication"], "aditivo": ["additive"],
    "digitalizacao": ["scanning", "scan"], "malha": ["mesh"],
    "escala": ["scale", "scaling"], "forca": ["force", "strength"],
    "punho": ["wrist"], "polegar": ["thumb"], "braco": ["arm"],
    "antropometrica": ["anthropometric"], "antropometricos": ["anthropometric"],
    "antropometricas": ["anthropometric"], "populacao": ["population"],
    "personalizacao": ["personaliz", "customiz"],
    "personalizada": ["personaliz", "customiz"],
    "participacao": ["participation"], "amostras": ["sample"],
    "inteligencia": ["intelligence"], "artificial": ["artificial"],
}


def tokens(s: str) -> set[str]:
    out = set()
    for raw in re.findall(r"[\w\-]{4,}", deaccent(s).lower()):
        for t in [raw] + PT_EN.get(raw, []):
            t = re.sub(r"(mente|coes|ares|ing|tion|tions|ity|ness|s)$", "", t)
            if len(t) >= 4 and t not in STOP:
                out.add(t)
    return out


def load_pdf(pdf_path: Path, cache: dict) -> tuple[list[str], str]:
    """Devolve (frases, diagnóstico de dificuldades de leitura)."""
    if pdf_path in cache:
        return cache[pdf_path]
    issues = []
    sents: list[str] = []
    try:
        doc = fitz.open(str(pdf_path))
        if doc.needs_pass:
            issues.append("protegido por palavra-passe")
        else:
            text = " ".join(page.get_text() for page in doc)
            n_pages = len(doc)
            per_page = len(text) / max(n_pages, 1)
            if per_page < 200:
                issues.append("sem camada de texto utilizável "
                              f"(~{int(per_page)} car./página — provável "
                              "digitalização sem OCR)")
            else:
                words = re.findall(r"[A-Za-zÀ-ÿ]{3,}", text)
                letter_ratio = (sum(len(w) for w in words) / len(text)
                                if text else 0)
                if letter_ratio < 0.45:
                    issues.append("texto extraído com muito ruído "
                                  "(fórmulas/tabelas ou codificação)")
                sents = sentences(norm_txt(text))
                if len(sents) < 15:
                    issues.append(f"poucas frases extraíveis ({len(sents)})")
        doc.close()
    except Exception as e:
        issues.append(f"PDF não abre: {type(e).__name__}")
    cache[pdf_path] = (sents, "; ".join(issues))
    return cache[pdf_path]


def best_excerpt(claim: str, pdf_path: Path, cache: dict,
                 title: str = "") -> str:
    sents, _ = load_pdf(pdf_path, cache)
    if not sents:
        return ""
    want = tokens(claim)
    title_toks = tokens(title) if title else set()
    best, score = "", 0
    for s in sents:
        if not (40 <= len(s) <= 600):
            continue
        stoks = tokens(s)
        # ignora cabeçalhos/estampas que reproduzem o título do artigo
        if title_toks and len(title_toks & stoks) >= max(3, len(title_toks) * 0.6):
            continue
        sc = len(want & stoks)
        if sc > score:
            best, score = s, sc
    return best[:450] if score >= 2 else ""


def main() -> None:
    # ---------- DOCX ----------
    root = etree.fromstring(ZipFile(DOCX).read("word/document.xml"))
    paras = root.findall(".//" + W + "p")
    texts = [para_text(p) for p in paras]
    bib_start = next(i for i, t in enumerate(texts)
                     if t.strip().startswith("Akasaka"))
    bib_end = next(i for i in range(bib_start + 1, len(texts))
                   if re.match(r"Anexo\s+A", texts[i].strip()))

    bib = []
    for t in texts[bib_start:bib_end]:
        t = norm_txt(t)
        if len(t) > 40:
            parsed = parse_bib_entry(t)
            if parsed:
                bib.append(parsed)
    idx = defaultdict(list)
    for i, b in enumerate(bib):
        first = b["surnames"][0]
        idx[(key_name(first), b["year"])].append(i)
        low = deaccent(b["authors"]).lower()
        for acr, org in ORG_ALIASES.items():
            if low.startswith(org):
                idx[(acr, b["year"])].append(i)
        if low.startswith("parlamento europeu"):
            idx[("parlamento europeu", b["year"])].append(i)
        if "openscad project" in low:
            idx[("openscad project", b["year"])].append(i)
        if "openscad community" in low:
            idx[("openscad community", b["year"])].append(i)
        if low.startswith("design council"):
            idx[("design council", b["year"])].append(i)
        if low.startswith("center for universal design"):
            idx[("center for universal design", b["year"])].append(i)

    matches = json.load(open(MATCHES))
    def pdf_for(entry) -> Path | None:
        cand, best = None, 0.0
        et = set(deaccent(entry["title"]).lower().split())
        for m in matches:
            mt = set(deaccent(m["title"]).lower().split())
            if not mt:
                continue
            s = len(et & mt) / max(len(et), len(mt))
            if s > best:
                cand, best = m, s
        if cand and best >= 0.6:
            p = ROOT / cand["pdf"]
            return p if p.exists() else None
        return None

    # ---------- PDF (paginação) ----------
    pdf = fitz.open(str(PDF))
    page_norm = [re.sub(r"\s+", " ", p.get_text()) for p in pdf]

    def page_of(fragment: str):
        frag = norm_txt(fragment)[:80]
        for i, pt in enumerate(page_norm):
            if frag in pt:
                n = i + 1 - PRELIM_OFFSET
                return n if n > 0 else f"prelim. (PDF {i+1})"
        return "?"

    # ---------- ocorrências ----------
    def resolve(first: str, second, year: str, seg: str):
        low = deaccent(first).lower()
        for org_key in ("parlamento europeu", "openscad project",
                        "openscad community", "design council",
                        "center for universal design"):
            if low.startswith(org_key):
                k = (org_key, year[:4] if year[0].isdigit() else "n.d.")
                break
        else:
            base = "n.d." if year.startswith("n.d") else year[:4]
            k = (key_name(first) if first not in ("ISO", "IEC", "ASTM")
                 else first.lower(), base)
        cand = idx.get(k, [])
        if len(cand) > 1:
            inits = initials_of(first)
            if inits:
                f = [i for i in cand
                     if deaccent(bib[i]["authors"]).replace(" ", "")
                     .startswith(deaccent(re.sub(r"[A-ZÀ-Þ]\.\s*", "", first))
                                 .strip().replace(" ", "") + ",")
                     and inits[0] in re.sub(r"[a-zà-þ,.\s]", "",
                                            bib[i]["authors"].split("(")[0])]
                # desambiguação simples pela inicial do primeiro autor
                g = []
                for i in cand:
                    m = re.search(r",\s*([A-ZÀ-Þ])\.", bib[i]["authors"])
                    if m and m.group(1) == inits[-1]:
                        g.append(i)
                if len(g) == 1:
                    cand = g
            if second and len(cand) > 1:
                f = [i for i in cand if len(bib[i]["surnames"]) >= 2
                     and key_name(bib[i]["surnames"][1]) == key_name(second)]
                if f:
                    cand = f
            if "et al" in seg and len(cand) > 1:
                f = [i for i in cand if len(bib[i]["surnames"]) >= 3]
                if f:
                    cand = f
        return cand

    rows = []
    excerpt_cache: dict = {}
    seen_pairs = set()
    for pi, p in enumerate(paras):
        if bib_start <= pi < bib_end:
            continue
        text = texts[pi]
        if not text.strip():
            continue
        if any("PAGEREF" in (n.text or "") for n in p.iter(W + "instrText")):
            continue
        occs = []
        for m in re.finditer(r"\(([^()]{4,300}?)\)", text):
            for seg in m.group(1).split(";"):
                sm = SEGMENT.match(seg.strip())
                if sm:
                    occs.append((m.start(), seg.strip(), sm.group(1),
                                 sm.group(2), sm.group(3)))
        for m in NARRATIVE.finditer(text):
            if text[max(0, m.start() - 1):m.start()] == "(":
                continue
            occs.append((m.start(), text[m.start():m.end()], m.group(1),
                         m.group(2), m.group(3)))
        for pos, seg, first, second, year in occs:
            claim = claim_sentence(text, pos)
            cand = resolve(first, second, year, seg)
            if not cand:
                rows.append({"page": page_of(text), "claim": claim,
                             "title": f"[sem entrada na bibliografia: {seg}]",
                             "authors": "", "year": year, "excerpt": "",
                             "conf": "sem correspondência",
                             "pdf_issues": ""})
                continue
            for i in cand[:1] if len(cand) == 1 else cand:
                b = bib[i]
                pair = (i, claim[:120])
                if pair in seen_pairs:
                    continue
                seen_pairs.add(pair)
                pdfp = pdf_for(b)
                if pdfp:
                    ex = best_excerpt(claim, pdfp, excerpt_cache, b["title"])
                    _, pdf_issues = load_pdf(pdfp, excerpt_cache)
                    conf = ("automático — verificar" if ex
                            else "PDF local sem excerto óbvio")
                else:
                    ex = ""
                    conf = "sem PDF local"
                    pdf_issues = "PDF não localizado no acervo"
                note = "" if len(cand) == 1 else " [AMBÍGUO]"
                rows.append({"page": page_of(text), "claim": claim,
                             "title": b["title"] + note,
                             "authors": b["authors"], "year": b["year"],
                             "excerpt": ex, "conf": conf,
                             "pdf_issues": pdf_issues or "nenhuma"})

    # ---------- XLSX ----------
    wb = Workbook()
    ws = wb.active
    ws.title = "Rastreabilidade"
    headers = ["Página DOCX", "Texto do DOCX (afirmação citante)",
               "Título do paper", "Autores", "Ano",
               "Excerto do paper (fundamentação)", "Confiança",
               "Dificuldades de leitura do PDF"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)
    for r in rows:
        ws.append([r["page"], r["claim"], r["title"], r["authors"],
                   r["year"], r["excerpt"], r["conf"], r["pdf_issues"]])
    widths = [10, 60, 45, 35, 8, 60, 22, 30]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = wrap
    ws.freeze_panes = "A2"
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    n_auto = sum(1 for r in rows if r["conf"].startswith("automático"))
    print(f"{len(rows)} linhas ({n_auto} com excerto automático) -> {OUT}")


if __name__ == "__main__":
    main()
